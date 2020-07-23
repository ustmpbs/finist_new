# -*- coding: utf-8 -*-
"""
Created on Wed Jul 22 13:57:14 2020

@author: Davit
"""

import openpyxl as opx
import os 
import pandas as pd
import numpy as np

os.chdir(r"D:\Bank of Russia\finist_new")

wb = opx.load_workbook("finist_new.xlsm")

NCalc = opx.wb['M'].cell(row = 2, cell = 2)

def calc_variations(wsheet):
    
    for Calc in range(1, NCalc):
        
        Application.Calculation = xlManual
        wb['Param'].cell(row = 2, column =1).value = wb['M'].cell(row =Calc+3, column = 2).value 
        wb['Param'].cell(row = 2, column =23).value = wb['M'].cell(row =Calc+3, column = 3).value 
        wb['Param'].cell(row = 2, column =24).value = wb['M'].cell(row =Calc+3, column = 4).value
        
        Balancing_all(wsheet)
        B_ALL_2 = wb.copy_worksheet('B_ALL')
        ws = wb['B_ALL_2'].active
        ws.title = wb['M'].cell(row = Calc +3, 5).value
        # фиксация результатов расчета
        
    