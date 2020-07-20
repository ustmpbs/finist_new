# -*- coding: utf-8 -*-
"""
Created on Mon Jul 20 08:12:29 2020

@author: Davit
"""

import openpyxl as opx
import os 
import pandas as pd
import numpy as np
import xlwings as xlw

os.chdir(r'D:\Bank of Russia\finist_new')

MaxT = 20
StartT = 3
ShiftT = 7
Bal_Result = 46
Bal_Rows = 81

delta = 0.1
    


# function Balancing_Bank_Semafori() as in Bal_mini
def Balancing_Bank_Semafori(wsheet):
    
    # очистка зоны балансировки
    for row in range(Bal_Result + 1, Bal_rows):
        for cell in range(StartT + ShiftT, MaxT + ShiftT):
            
            wsheet.cell(row = row, column = cell).value = None
    
    # фиксация исходного дисбаланса для оценки достаточности автоматической балансировки
    for t in range(StartT, MaxT):
        wsheet.cell(row = 4, column = ShiftT + t).value = wsheet.cell(row = 2, column = ShiftT + t).value
        
    # для каждого периода расчета = квартала
    
    for t in range(StartT, MaxT):
        
        Worksheets(ActiveSheet.Name).Calculate # придумать как это в питоне реализовать
        
        Pre_Bal = wsheet.cell(row = 2, column = ShiftT + t).value
        
        # Итерационно пока дисбаланс больше заданного уровня
        i =0 
        Lim_Res = wsheet.cell(row = 9, column = ShiftT + t).value
        Lim_Foreign = wsheet.cell(row = 10, column = ShiftT + t).value
        Lim_CBR = wsheet.cell(row = 11, column = ShiftT + t).value
        Lim_Repo = wsheet.cell(row = 12, column = ShiftT + t).value
        New_Loans = wsheet.cell(row = 38, column = ShiftT + t).value
        New_Deposits = wsheet.cell(row = 45, column = ShiftT + t).value
        
        
        while Pre_Bal > delta:
            
            if Pre_Bal < 0:
                
                # 1) увеличиваем ВЛА на запас до Макс лимита (но не более остаточного дисбаланса)
                Pre_Bal_m = min(wsheet.cell(row = 8, column = ShiftT + t).value, abs(Pre_Bal))
                
                for p in range(1, 2):
                    wsheet.cell(row = Bal_Result + p, column = ShiftT + t).value += Pre_Bal_m *  wsheet.cell(row = 12 + p, column = ShiftT + t).value
                Pre_Bal += Pre_Bal_m
                
                # 2) если нет дефицита капитала, закрываем нереализованный спрос на кредиты с прошлого квартала
                
                if Pre_Bal < delta and wsheet.cell(row = 6, column = ShiftT + t).value == 0 and wsheet.cell(row = Bal_Result + 19, column = ShiftT + t-1).value > 0:
                    Pre_Bal_m = min(abs(Pre_Bal), wsheet.cell(row = Bal_Result + 19, column = ShiftT + t-1).value)
                    
                    for p in range(1, 16): # по всем портфелям пропорционально их выдачам
                        wsheet.cell(row = Bal_Result + 2 + p, column = ShiftT + t).value += wsheet.cell(row = 21 + p, column = ShiftT + t).value * Pre_Bal_m / wsheet.cell(row = 38, column = ShiftT + t).value
                    Pre_Bal += Pre_Bal_m
                    wsheet.cell(row = Bal_Result + 19, column = ShiftT + t).value -= Pre_Bal_m
                    
                # 3) сокращаем новые выдачи депозитов и переносим в неотложенный спрос
                
                if Pre_Bal < delta and New_Deposits > 0:
                    Pre_Bal_m = min(abs(Pre_Bal), New_Deposits)
                    
                    for p in range(1, 6): # по всем портфелям пропорционально их выдачам
                        wsheet.cell(row = Bal_Result + 19 + p, column = ShiftT + t).value -= wsheet.cell(row = 38, column = ShiftT + t).value * Pre_Bal_m /  wsheet.cell(row = 45, column = ShiftT + t).value
                    Pre_Bal += Pre_Bal_m
                    wsheet.cell(row = Bal_Result + 26, column = ShiftT + t).value += Pre_Bal_m
                    
                # 4) затем покупаем облигации (FV для возможности продаж, по структуре FV+OCI)
                
                if Pre_Bal < delta:
                    
                    # в зависимости от наличия дефицита капитала во все бумаги
                    if wsheet.cell(row = 6, column = ShiftT + t).value == 0 :
                        for p in range(1, 6): # во все портфели пропорционально  объемам Т-1
                            wsheet.cell(row = Bal_Result + 26 + p, column = ShiftT + t).value += wsheet.cell(row = 14 + p, column = ShiftT + t).value + Pre_Bal
                    else:
                        for p in range(1, 2): # или в только гос - 2 портфеля
                            wsheet.cell(row = Bal_Result + 26 + p, column = ShiftT + t).value += Pre_Bal * wsheet.cell(row = 14 + p, column = ShiftT + t).value /(wsheet.cell(row = 15, column = ShiftT + t).value + wsheet.cell(row = 16, column = ShiftT + t).value)
                
        
        
        
            else: # если отток
                
                # 1) сокращаем ВЛА на запас до лимита (но не более остаточного дисбаланса)
                Pre_Bal_m = min(Pre_Bal, wsheet.cell(row = 7, column = ShiftT + t).value)
                for p in range(1, 2):
                    wsheet.cell(row = Bal_Result + p, column =  ShiftT + t).value -= Pre_Bal_m * wsheet.cell(row = 12 + p, column =  ShiftT + t).value
                Pre_Bal = Pre_Bal - Pre_Bal_m
                
                # 2) закрываем нереализованный спрос на кредиты
                if Pre_Bal < delta and wsheet.cell(row = Bal_Result + 26, column = ShiftT + t).value > 0:
                    Pre_Bal_m = min(abs(Pre_Bal_m), wsheet.cell(row = Bal_Result + 26, column = ShiftT + t-1).value)
                    for p in range(1, 6): # по всем портфелям пропорционально их выдачам
                        wsheet.cell(row = Bal_Result + 19 + p, column = ShiftT + t).value += Pre_Bal_m *  wsheet.cell(row = 38 + p, column = ShiftT + t).value / wsheet.cell(row = 45, column = ShiftT + t).value
                    Pre_Bal = Pre_Bal - Pre_Bal_m
                    wsheet.cell(row = Bal_Result + 26, column = ShiftT + t-1).value -= Pre_Bal_m
                    
                # 3) привлекаем по лимитам 
                #   сначала РЕПО - минимум лимита и дисбаланса - 50:50 рубли и валюта
                if Pre_Bal > delta and Lim_Repo > 0:
                    Pre_Bal_m = min(Lim_Repo, Pre_Bal)
                    
                    wsheet.cell(row = Bal_Result + 33, column = ShiftT + t).value += Pre_Bal_m/2
                    wsheet.cell(row = Bal_Result + 34, column = ShiftT + t).value += Pre_Bal_m/2
                    
                    Lim_Repo -= Pre_Bal_m
                    Pre_Bal -= Pre_Bal_m
                    
                # затем смотрим групповые лимиты по резидентам - минимум дисбаланса и лимита                
                if Pre_Bal > delta and Lim_Res > 0:
                    Pre_Bal_m = min(Pre_Bal, Lim_Res)
                    wsheet.cell(row = Bal_Result + 33, column = ShiftT + t).value += Pre_Bal_m
                    Lim_Res -= Pre_Bal_m
                    Pre_Bal -= Pre_Bal_m
                 
                # затем - групповые лимиты по нерезидентам - минимум лимита и дисбаланса (его остатка)
                if Pre_Bal > delta and Lim_Foreign > 0:
                    Pre_Bal_m = min(Lim_Foreign, Pre_Bal)
                    
                    wsheet.cell(row = Bal_Result + 34, column = ShiftT + t).value += Pre_Bal_m
                    Lim_Foreign -= Pre_Bal_m
                    Pre_Bal -= Pre_Bal_m
                    
                # потом - лимит ЦБ - миниму лимита и дисбаланса (его остатка)
                if Pre_Bal > delta and Lim_CBR > 0:
                    Pre_Bal_m = min(Pre_Bal, Lim_CBR)                    
                    wsheet.cell(row = Bal_Result + 35, column = ShiftT + t).value += Pre_Bal_m
                    
                    Lim_CBR -= Pre_Bal_m
                    Pre_Bal -= Pre_Bal_m
                    
                # 4) сокращаем выдачи кредитов
                if Pre_Bal > delta and New_loans > 0:
                    Pre_Bal_m = min(Pre_Bal, New_loans)
                    for p in range(1, 16): # пропорционально их выдачам 
                        wsheet.cell(row = Bal_Result + 2+p, column = ShiftT + t).value -= wsheet.cell(row = 21+p, column = ShiftT + t).value * Pre_Bal_m / wsheet.cell(row = 38, column = ShiftT + t).value
                    Pre_Bal -= Pre_Bal_m
                    New_Loans -= Pre_Bal_m
                    
                    # фиксируем отложенный спрос
                    wsheet.cell(row = Bal_Result + 19, column = ShiftT + t).value += Pre_Bal_m
                    
                # 5) последний шаг - финансирование в статусе тех дефолта
                if Pre_Bal > delta:
                    wsheet.cell(row = Bal_Result + 37, column = ShiftT + t).value = 1
                    wsheet.cell(row = Bal_Result + 35, column = ShiftT + t).value += Pre_Bal_m
            
            Worksheets(ActiveSheet.Name).Calculate
            Pre_Bal = wsheet.cell(row = 2, column = ShiftT + t).value
            i += 1
                    
                
                
                    
                
            
        
        
        
        
        
        
    
    
